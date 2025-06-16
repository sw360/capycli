# -------------------------------------------------------------------------------
# Copyright (c) 2019-24 Siemens
# All Rights Reserved.
# Author: thomas.graf@siemens.com
#
# SPDX-License-Identifier: MIT
# -------------------------------------------------------------------------------

import os
import sys
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import capycli.common.html_support
import capycli.common.json_support
import capycli.common.script_base
from capycli import get_logger
from capycli.bom.map_bom import MapBom
from capycli.common.map_result import MapResult
from capycli.common.print import print_red, print_text
from capycli.main.result_codes import ResultCode

LOG = get_logger(__name__)


class MappingToExcelXlsx(capycli.common.script_base.ScriptBase):
    """Create an Excel sheet showing the mapping result"""

    def define_styles(self) -> None:
        """Define style for Excel"""
        # fontDefault = Font(
        #    name="Calibri", size=11, bold=False, italic=False,vertAlign=None,
        #    underline="none", strike=False, color="FF000000")
        # fillDefault = PatternFill(
        #    fill_type=None, start_color="FFFFFFFF", end_color="FF000000")
        # borderDefault = Border(
        #    left=Side(border_style=None, color="FF000000"),
        #    right=Side(border_style=None, color="FF000000"),
        #    top=Side(border_style=None, color="FF000000"),
        #    bottom=Side(border_style=None, color="FF000000"),
        #    diagonal=Side(border_style=None, color="FF000000"),
        #    diagonal_direction=0,
        #    outline=Side(border_style=None, color="FF000000"),
        #    vertical=Side(border_style=None, color="FF000000"),
        #    horizontal=Side(border_style=None, color="FF000000"))

        self.fontBold = Font(name="Calibri", size=11, bold=True)
        self.fontBoldBlue = Font(name="Calibri", size=11, bold=True, color="FF0000FF")
        self.fontBoldOrange = Font(name="Calibri", size=11, bold=True, color="FFFFCC00")
        self.fontBoldRed = Font(name="Calibri", size=11, bold=True, color="FFFF0000")
        self.fontBold16 = Font(name="Calibri", size=16, bold=True)

        self.fillGray = PatternFill(
            fill_type="solid", start_color="FFC0C0C0", end_color="FFC0C0C0")
        self.border = Border(
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
            bottom=Side(border_style="thin", color="FF000000"))

    def mapping_result_to_xlsx(self, details: List[Dict[str, Any]], outputfile: str) -> None:
        """Create an Excel sheet showing the mapping overview"""
        wb = Workbook()
        ws = wb.active
        if not ws:
            return

        self.define_styles()

        ws["A1"] = "Mapping Result Overview"
        ws["A1"].font = self.fontBold16

        ws["A3"] = "SBOM Component"
        ws["A3"].fill = self.fillGray
        ws["A3"].border = self.border
        ws["B3"] = "Mapping Result"
        ws["B3"].fill = self.fillGray
        ws["B3"].border = self.border
        ws["C3"] = "Matching Component"
        ws["C3"].fill = self.fillGray
        ws["C3"].border = self.border

        row = 4
        max_width_component = 0
        max_width_result = 0
        max_width_match = 0
        for mapresult in details:
            versiontext = mapresult["BomItem"]["Name"]
            if "Version" in mapresult["BomItem"]:
                versiontext = versiontext + ", " + mapresult["BomItem"]["Version"]

            if (mapresult["Result"] == MapResult.INVALID) or \
                    (mapresult["Result"] == MapResult.NO_MATCH):
                font = self.fontBoldRed
            else:
                if MapBom.is_good_match(mapresult["Result"]):
                    font = self.fontBoldBlue
                else:
                    font = self.fontBoldOrange

            c = ws.cell(row=row, column=1, value=versiontext)
            c.font = font
            c.border = self.border
            c.alignment = Alignment(vertical="top")
            if len(versiontext) > max_width_component:
                max_width_component = len(versiontext)

            text = MapResult.map_code_to_string(mapresult["Result"]) + \
                " (" + str(mapresult["Result"]) + ")"
            c = ws.cell(row=row, column=2, value=text)
            c.font = font
            c.border = self.border
            c.alignment = Alignment(vertical="top")
            if len(text) > max_width_result:
                max_width_result = len(text)

            text = ""
            if (mapresult["Result"] == MapResult.INVALID) or \
                    (mapresult["Result"] == MapResult.NO_MATCH):
                text = "(none)"
            else:
                for matchitem in mapresult["Matches"]:
                    line = matchitem["Name"] + ", " + matchitem["Version"] + "\n"
                    if len(line) > max_width_match:
                        max_width_match = len(line)

                    text += line
                    mid = ""
                    if "RepositoryType" in matchitem:
                        mid = matchitem["RepositoryType"]
                    if "RepositoryId" in matchitem:
                        mid = mid + " = " + matchitem["RepositoryId"]

                    if mid:
                        text = mid + "\n"

                    if len(line) > max_width_match:
                        max_width_match = len(line)

            c = ws.cell(row=row, column=3, value=text)
            c.alignment = Alignment(wrapText=True)
            c.font = font
            c.border = self.border

            row += 1

        ws.column_dimensions["A"].width = max_width_component
        ws.column_dimensions["B"].width = max_width_result
        ws.column_dimensions["C"].width = max_width_match

        try:
            wb.save(outputfile)
        except Exception as ex:
            print_red("Error writing Excel sheet: " + repr(ex))
            sys.exit(ResultCode.RESULT_ERROR_WRITING_FILE)

    def run(self, args: Any) -> None:
        """Main method()"""
        if args.debug:
            global LOG
            LOG = capycli.get_logger(__name__)

        print_text(
            "\n" + capycli.get_app_signature() +
            " - Create an Excel sheet showing the mapping result\n")

        if args.help:
            print("usage: CaPyCli mapping toxlsx -i <mapping_result.json> -o <mapping_result.xlsx>")
            print("")
            print("optional arguments:")
            print("-h, --help            show this help message and exit")
            print("-i INPUTFILE,         input mapping result JSON file to read from")
            print("-o OUTPUTFILE,        output XLSX file to write to")
            print("")
            return

        if not args.inputfile:
            print_red("No input file specified!")
            sys.exit(ResultCode.RESULT_COMMAND_ERROR)

        if not os.path.isfile(args.inputfile):
            print_red("Input file not found!")
            sys.exit(ResultCode.RESULT_FILE_NOT_FOUND)

        if not args.outputfile:
            print_red("No output file specified!")
            sys.exit(ResultCode.RESULT_COMMAND_ERROR)

        print_text("Loading mapping result " + args.inputfile)
        try:
            mapping_result = capycli.common.json_support.load_json_file(args.inputfile)
        except Exception as ex:
            print_red("Error reading input file: " + repr(ex))
            sys.exit(ResultCode.RESULT_COMMAND_ERROR)

        print_text("Creating Excel sheet " + args.outputfile)
        self.mapping_result_to_xlsx(mapping_result, args.outputfile)

        print()
