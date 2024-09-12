# -------------------------------------------------------------------------------
# Copyright (c) 2019-2024 Siemens
# All Rights Reserved.
# Author: thomas.graf@siemens.com
#
# SPDX-License-Identifier: MIT
# -------------------------------------------------------------------------------

import os
import sys
from typing import Any, Dict

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

import capycli.common.html_support
import capycli.common.json_support
import capycli.common.script_base
from capycli import get_logger
from capycli.bom.map_bom import MapBom
from capycli.common.map_result import MapResult
from capycli.common.print import print_red, print_text
from capycli.main.result_codes import ResultCode

LOG = get_logger(__name__)


class MappingOverviewToExcelXlsx(capycli.common.script_base.ScriptBase):
    """Create an Excel sheet showing the mapping overview."""

    def define_styles(self) -> None:
        """Define style for Excel"""
        # fontDefault = Font(
        #    name="Calibri", size=11, bold=False, italic=False,vertAlign=None,
        #    underline="none", strike=False, color="FF000000")
        # fillDefault = PatternFill(
        #    fill_type=None, start_color="FFFFFFFF", end_color="FF000000")
        # borderDefault = Border(
        #    left=Side(border_style=None, color="FF000000"),
        #    right=Side(border_style=None, color="FF000000"),
        #    top=Side(border_style=None, color="FF000000"),
        #    bottom=Side(border_style=None, color="FF000000"),
        #    diagonal=Side(border_style=None, color="FF000000"),
        #    diagonal_direction=0,
        #    outline=Side(border_style=None, color="FF000000"),
        #    vertical=Side(border_style=None, color="FF000000"),
        #    horizontal=Side(border_style=None, color="FF000000"))

        self.fontBold = Font(name="Calibri", size=11, bold=True)
        self.fontBoldBlue = Font(name="Calibri", size=11, bold=True, color="FF0000FF")
        self.fontBoldOrange = Font(name="Calibri", size=11, bold=True, color="FFFFCC00")
        self.fontBoldRed = Font(name="Calibri", size=11, bold=True, color="FFFF0000")
        self.fontBold16 = Font(name="Calibri", size=16, bold=True)

        self.fillGray = PatternFill(
            fill_type="solid", start_color="FFC0C0C0", end_color="FFC0C0C0")
        self.border = Border(
            left=Side(border_style="thin", color="FF000000"),
            right=Side(border_style="thin", color="FF000000"),
            top=Side(border_style="thin", color="FF000000"),
            bottom=Side(border_style="thin", color="FF000000"))

    def mapping_overview_to_xlsx(self, overview: Dict[str, Any], outputfile: str) -> None:
        """Create an Excel sheet showing the mapping overview"""
        wb = Workbook()
        ws = wb.active
        if not ws:
            return

        self.define_styles()

        ws["A1"] = "Mapping Result Overview"
        ws["A1"].font = self.fontBold16

        ws["A3"] = "Overall Result: " + overview["OverallResult"]
        if overview["OverallResult"] == "COMPLETE":
            ws["A3"].font = self.fontBoldBlue
        else:
            ws["A3"].font = self.fontBoldRed

        ws["A5"] = "Component"
        ws["A5"].fill = self.fillGray
        ws["A5"].border = self.border
        ws["B5"] = "Mapping Result"
        ws["B5"].fill = self.fillGray
        ws["B5"].border = self.border
        ws["C5"] = "Mapping Code"
        ws["C5"].fill = self.fillGray
        ws["C5"].border = self.border

        row = 6
        max_width_component = 0
        max_width_text = 0
        for item in overview["Details"]:
            if (item["ResultCode"] == MapResult.INVALID) or \
                    (item["ResultCode"] == MapResult.NO_MATCH):
                font = self.fontBoldRed
            else:
                if MapBom.is_good_match(item["ResultCode"]):
                    font = self.fontBoldBlue
                else:
                    font = self.fontBoldOrange

            c = ws.cell(row=row, column=1, value=item["BomItem"])
            c.font = font
            c.border = self.border
            if len(item["BomItem"]) > max_width_component:
                max_width_component = len(item["BomItem"])

            c = ws.cell(row=row, column=2, value=item["ResultText"])
            c.font = font
            c.border = self.border
            if len(item["ResultText"]) > max_width_text:
                max_width_text = len(item["ResultText"])

            c = ws.cell(row=row, column=3, value=item["ResultCode"])
            c.font = font
            c.border = self.border

            row += 1

        ws.column_dimensions["A"].width = max_width_component
        ws.column_dimensions["B"].width = max_width_text
        ws.column_dimensions["C"].width = len("Mapping Code")

        wb.save(outputfile)

    def run(self, args: Any) -> None:
        """Main method()"""
        if args.debug:
            global LOG
            LOG = capycli.get_logger(__name__)

        print_text(
            "\n" + capycli.APP_NAME + ", " + capycli.get_app_version() +
            " - Create an Excel sheet showing the mapping overview\n")

        if args.help:
            print("usage: CaPyCli moverview toxlsx -i <mapping_overview.json> -o <mapping_overview.xlsx>")
            print("")
            print("optional arguments:")
            print("-h, --help            show this help message and exit")
            print("-i INPUTFILE,         input mapping result JSON file to read from")
            print("-o OUTPUTFILE,        output XLSX file to write to")
            print("")
            return

        if not args.inputfile:
            print_red("No input file specified!")
            sys.exit(ResultCode.RESULT_COMMAND_ERROR)

        if not os.path.isfile(args.inputfile):
            print_red("Input file not found!")
            sys.exit(ResultCode.RESULT_FILE_NOT_FOUND)

        if not args.outputfile:
            print_red("No output file specified!")
            sys.exit(ResultCode.RESULT_COMMAND_ERROR)

        print_text("Loading mapping overview " + args.inputfile)
        try:
            overview = capycli.common.json_support.load_json_file(args.inputfile)
        except Exception as ex:
            print_red("Error reading input file: " + repr(ex))
            sys.exit(ResultCode.RESULT_COMMAND_ERROR)

        print_text("Creating Excel sheet " + args.outputfile)
        try:
            self.mapping_overview_to_xlsx(overview, args.outputfile)
        except Exception as ex:
            print_red("Error creating overview file: " + repr(ex))
            sys.exit(ResultCode.RESULT_GENERAL_ERROR)

        print()
